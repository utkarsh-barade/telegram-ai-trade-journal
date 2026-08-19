"""
Unit tests for the Excel export service.
"""

from __future__ import annotations

import io
from datetime import datetime

import openpyxl
import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from db.models import Base
from parser.trade_parser import ParseResult
from services.export_service import export_trades_to_excel
from services.trade_service import create_trade


# ──────────────────────────────────────────────────────────────────────────────
# Fixtures
# ──────────────────────────────────────────────────────────────────────────────

@pytest.fixture
def db():
    engine = create_engine("sqlite:///:memory:", connect_args={"check_same_thread": False})
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)
    session = Session()
    yield session
    session.close()


def _make_parse(stock="DLF", direction="BUY") -> ParseResult:
    return ParseResult(
        stock=stock,
        strike=650.0,
        option_type="CE",
        direction=direction,
        entry_price=24.0,
        stop_loss=22.0,
        target=27.0,
        expiry="Aug 2026",
        trade_date=datetime(2026, 8, 15),
        date_is_explicit=True,
        instrument=f"{stock} 650 CE",
        raw_text=f"{stock} 650 CE BUY @24 SL22 TG27",
        missing_fields=[],
        is_complete=True,
    )


# ──────────────────────────────────────────────────────────────────────────────
# Export tests
# ──────────────────────────────────────────────────────────────────────────────

class TestExcelExport:
    def test_returns_bytes_io(self, db):
        buf = export_trades_to_excel(db, user_id=123)
        assert isinstance(buf, io.BytesIO)

    def test_valid_xlsx(self, db):
        create_trade(db, _make_parse(), user_id=123)
        buf = export_trades_to_excel(db, user_id=123)
        wb = openpyxl.load_workbook(buf)
        assert wb is not None

    def test_sheet_name(self, db):
        buf = export_trades_to_excel(db)
        wb = openpyxl.load_workbook(buf)
        assert "Trade Journal" in wb.sheetnames

    def test_header_row(self, db):
        buf = export_trades_to_excel(db)
        wb = openpyxl.load_workbook(buf)
        ws = wb["Trade Journal"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, 10)]
        assert "Trade ID" in headers
        assert "Stock" in headers

    def test_data_row_written(self, db):
        create_trade(db, _make_parse(), user_id=123)
        buf = export_trades_to_excel(db, user_id=123)
        wb = openpyxl.load_workbook(buf)
        ws = wb["Trade Journal"]
        # Row 2 should have data (row 1 is header)
        values = [ws.cell(row=2, column=c).value for c in range(1, 5)]
        assert any(v is not None for v in values)

    def test_empty_export_has_only_header(self, db):
        buf = export_trades_to_excel(db, user_id=999)
        wb = openpyxl.load_workbook(buf)
        ws = wb["Trade Journal"]
        assert ws.max_row == 1  # only header

    def test_multiple_trades(self, db):
        create_trade(db, _make_parse("DLF"), user_id=123)
        create_trade(db, _make_parse("NIFTY"), user_id=123)
        buf = export_trades_to_excel(db, user_id=123)
        wb = openpyxl.load_workbook(buf)
        ws = wb["Trade Journal"]
        assert ws.max_row == 3  # header + 2 data rows
