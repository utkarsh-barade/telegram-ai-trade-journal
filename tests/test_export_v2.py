"""
Unit tests for 4-Sheet Excel Export (services/export_service.py).
"""

from __future__ import annotations

import openpyxl
import pytest
from datetime import datetime
from db.models import Direction, OptionType, Trade, TradeOutcome, TradeTarget, TradeTargetStatus
from services.export_service import export_trades_to_excel_v2
from services.analytics_service import TradeFilter


@pytest.fixture
def export_db(db):
    t1 = Trade(
        stock="RELIANCE",
        instrument="RELIANCE 3000 CE",
        strike=3000.0,
        option_type=OptionType.CE,
        direction=Direction.BUY,
        entry_price=100.0,
        stop_loss=90.0,
        target=120.0,
        weighted_exit_price=115.0,
        remaining_qty_pct=0.0,
        outcome=TradeOutcome.WIN,
        pnl_inr=3000.0,
        capital=100000.0,
        analyst_id=1,
        analyst_username="analyst_test",
        trade_date=datetime(2026, 8, 15, 9, 30),
        raw_message="RELIANCE 3000 CE @100",
    )
    db.add(t1)
    db.flush()

    leg1 = TradeTarget(trade_id=t1.id, level="TG1", target_price=110.0, planned_qty_pct=50.0, status=TradeTargetStatus.HIT, exit_price=110.0)
    leg2 = TradeTarget(trade_id=t1.id, level="FINAL", target_price=120.0, planned_qty_pct=50.0, status=TradeTargetStatus.HIT, exit_price=120.0)
    db.add_all([leg1, leg2])
    db.commit()
    return db


class TestExportV2:
    def test_export_generates_4_sheets(self, export_db):
        buf = export_trades_to_excel_v2(export_db, filter_params=TradeFilter())
        wb = openpyxl.load_workbook(buf)

        sheet_names = wb.sheetnames
        assert len(sheet_names) == 4
        assert "Trade Journal" in sheet_names
        assert "Performance Summary" in sheet_names
        assert "Daily Summary" in sheet_names
        assert "Target Legs Detail" in sheet_names

    def test_sheet1_journal_columns(self, export_db):
        buf = export_trades_to_excel_v2(export_db, filter_params=TradeFilter())
        wb = openpyxl.load_workbook(buf)
        ws1 = wb["Trade Journal"]

        headers = [ws1.cell(row=1, column=col).value for col in range(1, ws1.max_column + 1)]
        assert "Trade ID" in headers
        assert "Weighted Exit ₹" in headers
        assert "Targets Summary" in headers
        assert ws1.cell(row=2, column=1).value == "#001"

    def test_sheet4_target_legs_detail(self, export_db):
        buf = export_trades_to_excel_v2(export_db, filter_params=TradeFilter())
        wb = openpyxl.load_workbook(buf)
        ws4 = wb["Target Legs Detail"]

        assert ws4.max_row == 3  # Header + 2 target legs
        assert ws4.cell(row=2, column=4).value == "TG1"
        assert ws4.cell(row=3, column=4).value == "FINAL"
