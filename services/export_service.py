"""
Excel export service — Phase 2 Multi-Sheet Filter-Aware Export.

Generates 4 worksheets:
  1. Trade Journal (Main journal with compact targets summary & weighted exit)
  2. Performance Summary (Analyst performance metrics, win rate, expectancy, drawdown)
  3. Daily Summary (Date-wise breakdown of trades, wins/losses, net P&L)
  4. Target Legs Detail (Line-by-line breakdown of all target legs)
"""

from __future__ import annotations

import io
from datetime import datetime
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from db.models import Trade, TradeOutcome
from services.analytics_service import TradeFilter, build_filtered_query, get_overview_metrics


# ──────────────────────────────────────────────────────────────────────────────
# Formatting helpers
# ──────────────────────────────────────────────────────────────────────────────

def _fmt_date(dt: Optional[datetime]) -> str:
    return dt.strftime("%d-%b-%Y") if dt else ""


def _fmt_datetime(dt: Optional[datetime]) -> str:
    return dt.strftime("%d-%b-%Y %H:%M") if dt else ""


def _fmt_time(dt: Optional[datetime]) -> str:
    return dt.strftime("%H:%M") if dt else ""


def _pct(val: Optional[float]) -> str:
    return f"{val:.2f}%" if val is not None else ""


def _targets_summary(t: Trade) -> str:
    """Build compact string e.g. 25.5(✓40%) / 26.5(✓30%) / 27(30%)."""
    if not t.targets:
        return f"{t.target}" if t.target else ""

    parts = []
    for leg in t.targets:
        check = "✓" if leg.status and leg.status.value == "HIT" else ""
        qty = int(leg.planned_qty_pct) if leg.planned_qty_pct == int(leg.planned_qty_pct) else leg.planned_qty_pct
        parts.append(f"{leg.target_price}({check}{qty}%)")
    return " / ".join(parts)


def _outcome_fill(outcome: str) -> Optional[PatternFill]:
    colors = {
        "WIN": "C6EFCE",           # light green
        "LOSS": "FFC7CE",          # light red
        "BREAKEVEN": "FFEB9C",     # light yellow
        "PARTIAL_EXIT": "FCE4D6",  # light orange
        "OPEN": "DDEBF7",          # light blue
        "NEEDS_REVIEW": "F4CCFF",    # light purple
    }
    hex_color = colors.get(outcome)
    return PatternFill("solid", fgColor=hex_color) if hex_color else None


# ──────────────────────────────────────────────────────────────────────────────
# Sheet 1 Columns
# ──────────────────────────────────────────────────────────────────────────────

_SHEET1_COLUMNS = [
    ("Trade ID",            lambda t: t.display_id),
    ("Trade Date",          lambda t: _fmt_date(t.trade_date)),
    ("Entry Time",          lambda t: _fmt_time(t.entry_time)),
    ("Stock",               lambda t: t.stock),
    ("Instrument",          lambda t: t.instrument_label),
    ("Strike",              lambda t: t.strike),
    ("Option Type",         lambda t: t.option_type.value if t.option_type else ""),
    ("Expiry",              lambda t: t.expiry or ""),
    ("Direction",           lambda t: t.direction.value if t.direction else ""),
    ("Entry ₹",             lambda t: t.entry_price),
    ("SL ₹",                lambda t: t.stop_loss),
    ("Targets Summary",     lambda t: _targets_summary(t)),
    ("Weighted Exit ₹",     lambda t: t.weighted_exit_price or t.exit_price or ""),
    ("Exit Date",           lambda t: _fmt_date(t.exit_datetime)),
    ("Exit Time",           lambda t: _fmt_time(t.exit_datetime)),
    ("Outcome",             lambda t: t.outcome.value if t.outcome else ""),
    ("P&L ₹",               lambda t: t.pnl_inr if t.pnl_inr is not None else ""),
    ("P&L %",               lambda t: _pct(t.pnl_pct)),
    ("Capital ₹",           lambda t: t.capital if t.capital else ""),
    ("Capital P&L %",       lambda t: _pct(t.capital_pnl_pct)),
    ("Risk ₹",              lambda t: t.risk_inr if t.risk_inr else ""),
    ("Risk %",              lambda t: _pct(t.risk_pct)),
    ("Planned R:R",         lambda t: t.planned_rr if t.planned_rr else ""),
    ("Achieved R:R",        lambda t: t.achieved_rr if t.achieved_rr else ""),
    ("Analyst",             lambda t: t.analyst_username or f"User #{t.analyst_id}"),
    ("Notes",               lambda t: t.notes or ""),
]


# ──────────────────────────────────────────────────────────────────────────────
# Sheet 4 Columns
# ──────────────────────────────────────────────────────────────────────────────

_SHEET4_COLUMNS = [
    ("Trade ID",       lambda t, leg: t.display_id),
    ("Stock",          lambda t, leg: t.stock),
    ("Instrument",     lambda t, leg: t.instrument_label),
    ("Level",          lambda t, leg: leg.level),
    ("Target Price ₹", lambda t, leg: leg.target_price),
    ("Qty %",          lambda t, leg: _pct(leg.planned_qty_pct)),
    ("Status",         lambda t, leg: leg.status.value if leg.status else ""),
    ("Exit Price ₹",   lambda t, leg: leg.exit_price if leg.exit_price is not None else ""),
    ("Exit DateTime",  lambda t, leg: _fmt_datetime(leg.exit_datetime)),
]


# ──────────────────────────────────────────────────────────────────────────────
# Main Export API
# ──────────────────────────────────────────────────────────────────────────────

def export_trades_to_excel_v2(
    db: Session,
    filter_params: Optional[TradeFilter] = None,
    user_id: Optional[int] = None,
) -> io.BytesIO:
    """
    Export trades to a 4-sheet Excel workbook based on active filter parameters.
    """
    if filter_params is None:
        filter_params = TradeFilter()

    if user_id is not None:
        filter_params.analyst_id = user_id

    trades = build_filtered_query(db, filter_params).order_by(Trade.trade_date.desc(), Trade.id.desc()).all()
    metrics = get_overview_metrics(db, filter_params)

    wb = openpyxl.Workbook()

    # Styling definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill_blue = PatternFill("solid", fgColor="1F4E79")
    header_fill_navy = PatternFill("solid", fgColor="2F5597")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── Sheet 1: Trade Journal ───────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Trade Journal"

    for col_idx, (header, _) in enumerate(_SHEET1_COLUMNS, start=1):
        cell = ws1.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill_blue
        cell.alignment = header_align

    ws1.row_dimensions[1].height = 30

    for row_idx, trade in enumerate(trades, start=2):
        for col_idx, (_, accessor) in enumerate(_SHEET1_COLUMNS, start=1):
            try:
                val = accessor(trade)
            except Exception:
                val = ""
            ws1.cell(row=row_idx, column=col_idx, value=val)

        outcome_val = trade.outcome.value if trade.outcome else ""
        fill = _outcome_fill(outcome_val)
        if fill:
            for col_idx in range(1, len(_SHEET1_COLUMNS) + 1):
                ws1.cell(row=row_idx, column=col_idx).fill = fill

    for col_idx in range(1, len(_SHEET1_COLUMNS) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max((len(str(ws1.cell(row=r, column=col_idx).value or "")) for r in range(1, ws1.max_row + 1)), default=10)
        ws1.column_dimensions[col_letter].width = min(max_len + 4, 45)
    ws1.freeze_panes = "A2"

    # ── Sheet 2: Performance Summary ─────────────────────────────────────────
    ws2 = wb.create_sheet(title="Performance Summary")
    ws2.cell(row=1, column=1, value="Performance Metric").font = header_font
    ws2.cell(row=1, column=1).fill = header_fill_navy
    ws2.cell(row=1, column=2, value="Value").font = header_font
    ws2.cell(row=1, column=2).fill = header_fill_navy
    ws2.row_dimensions[1].height = 30

    summary_data = [
        ("Total Trades", metrics["total_trades"]),
        ("Open Trades", metrics["open_trades"]),
        ("Wins", metrics["wins"]),
        ("Losses", metrics["losses"]),
        ("Breakevens", metrics["breakevens"]),
        ("Win Rate %", f"{metrics['win_rate']:.2f}%"),
        ("Net P&L ₹", f"₹{metrics['net_pnl']:,.2f}"),
        ("Capital Return %", f"{metrics['capital_return']:.2f}%"),
        ("Average P&L ₹", f"₹{metrics['avg_pnl']:,.2f}"),
        ("Average R:R", metrics["avg_rr"]),
        ("Profit Factor", metrics["profit_factor"]),
        ("Expectancy ₹", f"₹{metrics['expectancy']:,.2f}"),
        ("Maximum Drawdown ₹", f"₹{metrics['max_drawdown']:,.2f}"),
        ("Maximum Drawdown %", f"{metrics['max_drawdown_pct']:.2f}%"),
        ("Best Trade ₹", f"₹{metrics['best_trade']:,.2f}"),
        ("Worst Trade ₹", f"₹{metrics['worst_trade']:,.2f}"),
    ]

    for r_idx, (k, v) in enumerate(summary_data, start=2):
        cell_k = ws2.cell(row=r_idx, column=1, value=k)
        cell_k.font = Font(bold=True)
        ws2.cell(row=r_idx, column=2, value=v)

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 25
    ws2.freeze_panes = "A2"

    # ── Sheet 3: Daily Summary ───────────────────────────────────────────────
    ws3 = wb.create_sheet(title="Daily Summary")
    daily_headers = ["Date", "Number of Trades", "Wins", "Losses", "Win Rate %", "Net P&L ₹", "Capital Return %"]
    for col_idx, h in enumerate(daily_headers, start=1):
        cell = ws3.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill_blue
        cell.alignment = header_align
    ws3.row_dimensions[1].height = 30

    daily_dict: dict[str, dict[str, Any]] = {}
    for t in sorted(trades, key=lambda x: x.trade_date or datetime.min):
        date_str = _fmt_date(t.trade_date)
        if not date_str:
            continue
        if date_str not in daily_dict:
            daily_dict[date_str] = {"date": date_str, "trades": 0, "wins": 0, "losses": 0, "pnl": 0.0, "cap": t.capital}

        daily_dict[date_str]["trades"] += 1
        daily_dict[date_str]["pnl"] += (t.pnl_inr or 0.0)
        if t.outcome == TradeOutcome.WIN:
            daily_dict[date_str]["wins"] += 1
        elif t.outcome == TradeOutcome.LOSS:
            daily_dict[date_str]["losses"] += 1

    r_idx = 2
    for d_str, data in daily_dict.items():
        decided = data["wins"] + data["losses"]
        wr = f"{(data['wins'] / decided) * 100.0:.2f}%" if decided > 0 else "0.00%"
        cap = data["cap"]
        cap_ret = f"{(data['pnl'] / cap) * 100.0:.2f}%" if cap else "0.00%"

        ws3.cell(row=r_idx, column=1, value=data["date"])
        ws3.cell(row=r_idx, column=2, value=data["trades"])
        ws3.cell(row=r_idx, column=3, value=data["wins"])
        ws3.cell(row=r_idx, column=4, value=data["losses"])
        ws3.cell(row=r_idx, column=5, value=wr)
        ws3.cell(row=r_idx, column=6, value=data["pnl"])
        ws3.cell(row=r_idx, column=7, value=cap_ret)
        r_idx += 1

    for col_idx in range(1, len(daily_headers) + 1):
        col_letter = get_column_letter(col_idx)
        ws3.column_dimensions[col_letter].width = 22
    ws3.freeze_panes = "A2"

    # ── Sheet 4: Target Legs Detail ──────────────────────────────────────────
    ws4 = wb.create_sheet(title="Target Legs Detail")
    for col_idx, (header, _) in enumerate(_SHEET4_COLUMNS, start=1):
        cell = ws4.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill_navy
        cell.alignment = header_align
    ws4.row_dimensions[1].height = 30

    leg_row_idx = 2
    for trade in trades:
        if trade.targets:
            for leg in trade.targets:
                for col_idx, (_, accessor) in enumerate(_SHEET4_COLUMNS, start=1):
                    try:
                        val = accessor(trade, leg)
                    except Exception:
                        val = ""
                    ws4.cell(row=leg_row_idx, column=col_idx, value=val)
                leg_row_idx += 1

    for col_idx in range(1, len(_SHEET4_COLUMNS) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max((len(str(ws4.cell(row=r, column=col_idx).value or "")) for r in range(1, ws4.max_row + 1)), default=10)
        ws4.column_dimensions[col_letter].width = min(max_len + 4, 35)
    ws4.freeze_panes = "A2"

    # Save to BytesIO
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_trades_to_excel(
    db: Session,
    user_id: Optional[int] = None,
    filename_hint: str = "trade_journal",
) -> io.BytesIO:
    """Wrapper function for backward compatibility with Phase 1 export API."""
    return export_trades_to_excel_v2(db, user_id=user_id)
